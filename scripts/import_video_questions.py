from __future__ import annotations

import argparse
import json
import re
from datetime import date
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
OUTPUT_PATH = ROOT / "data" / "video-questions.json"
SECTION_SLUGS = {
    "1．情報社会": "information-society",
    "2．情報デザイン": "information-design",
    "3．デジタル": "digital",
    "4．ネットワーク": "network",
    "5．プログラミング": "programming",
}
# The source workbook leaves these two answer cells blank. Both answers are
# determined directly by their question statements and corresponding videos.
ANSWER_OVERRIDES = {
    84: "123457000000000000",
    88: "False",
}


def parse_number(value: object) -> int | None:
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    match = re.match(r"\s*0*(\d{1,3})(?:\D|$)", str(value or ""))
    return int(match.group(1)) if match else None


def split_keywords(value: object) -> list[str]:
    return [item.strip() for item in str(value or "").split(",") if item.strip()]


def read_videos(path: Path) -> dict[int, list[dict[str, str]]]:
    videos = json.loads(path.read_text(encoding="utf-8"))
    mapped: dict[int, list[dict[str, str]]] = {}
    for video in videos:
        number = parse_number(video.get("title"))
        video_id = str(video.get("id") or "").strip()
        title = str(video.get("title") or "").strip()
        if number is None or not video_id or not title:
            continue
        mapped.setdefault(number, []).append({"id": video_id, "title": title})
    return mapped


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Import public question fields and YouTube mappings without explanation text."
    )
    parser.add_argument("workbook", type=Path)
    parser.add_argument("channel_metadata", type=Path)
    parser.add_argument("--output", type=Path, default=OUTPUT_PATH)
    args = parser.parse_args()

    workbook_path = args.workbook.resolve()
    metadata_path = args.channel_metadata.resolve()
    if not workbook_path.is_file():
        raise SystemExit(f"Workbook not found: {workbook_path}")
    if not metadata_path.is_file():
        raise SystemExit(f"Channel metadata not found: {metadata_path}")

    videos_by_number = read_videos(metadata_path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=False)
    sections: list[dict[str, object]] = []
    seen_numbers: set[int] = set()

    for worksheet in workbook.worksheets:
        slug = SECTION_SLUGS.get(worksheet.title)
        if not slug:
            continue
        questions: list[dict[str, object]] = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            number = parse_number(row[0] if row else None)
            question = str(row[1] or "").strip() if len(row) > 1 else ""
            answer = str(row[2] or "").strip() if len(row) > 2 else ""
            keywords = split_keywords(row[4] if len(row) > 4 else None)
            if number is None or not question:
                continue
            if number in seen_numbers:
                raise SystemExit(f"Duplicate question number in workbook: {number}")
            answer = answer or ANSWER_OVERRIDES.get(number, "")
            if not answer:
                raise SystemExit(f"Question has no public answer: {number}")
            seen_numbers.add(number)
            questions.append(
                {
                    "number": number,
                    "question": question,
                    "answer": answer,
                    "keywords": keywords,
                    "videos": videos_by_number.get(number, []),
                }
            )
        sections.append(
            {
                "id": slug,
                "label": worksheet.title.split("．", 1)[-1],
                "questions": questions,
            }
        )

    expected_numbers = set(range(1, 331))
    missing_questions = sorted(expected_numbers - seen_numbers)
    extra_questions = sorted(seen_numbers - expected_numbers)
    missing_videos = sorted(number for number in seen_numbers if not videos_by_number.get(number))
    if missing_questions or extra_questions or missing_videos:
        raise SystemExit(
            "Import completeness check failed: "
            f"missing_questions={missing_questions} extra_questions={extra_questions} "
            f"missing_videos={missing_videos}"
        )

    payload = {
        "generated_on": date.today().isoformat(),
        "question_count": len(seen_numbers),
        "content_policy": "問題・答え・キーワード・対応動画のみ。解説本文は収録しない。",
        "youtube_channel": "https://www.youtube.com/@mei_chan_nel",
        "sections": sections,
    }
    output_path = args.output.resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(
        f"questions={len(seen_numbers)} sections={len(sections)} "
        f"videos={sum(len(videos_by_number[number]) for number in seen_numbers)} output={output_path}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
